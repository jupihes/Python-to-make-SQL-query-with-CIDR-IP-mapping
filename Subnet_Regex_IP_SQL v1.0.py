"""
@author: Hossein Khayami @https://github.com/h-khayami
@https://github.com/jupihes Hesam Mohammad Hosseini supported to revise and add all sections to make one unified program

"CIDR Notation to Regex Tool" http://d.xenowire.net/cidr2regex.php used and few bugs on ??? modified.

"""
import re
from netaddr import *
from openpyxl import load_workbook
import csv
from os import chdir
import pandas as pd

######## section 1: #######
# CIDR Notation to Regex Tool
# Original: http://d.xenowire.net/cidr2regex.php
# Converted to Python by 112buddyd
def cidr_to_regex(cidr):
   
    # Validation
    cidr_regex = r'^(([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\.){3}([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])(\/([0-9]|[1-2][0-9]|3[0-2]))$'
    if not re.match(cidr_regex, cidr):
        return 'Input not valid CIDR notation. Ex. 192.168.1.0/24'
    
    # Setup Regex string dictionary
    map = {}
    # 255
    map[0] = {}
    map[0][0] = '([0-9]|[1-9][0-9]|1[0-9][0-9]|2([0-4][0-9]|5[0-5]))'
    # 128
    map[1] = {}
    map[1][0] = '([0-9]|[1-9][0-9]|1[0-1][0-9]|12[0-7])'
    map[1][128] = '(12[8-9]|1[3-9][0-9]|2([0-4][0-9]|5[0-5]))'
    # 64
    map[2] = {}
    map[2][0] = '([0-9]|[1-5][0-9]|6[0-3])'
    map[2][64] = '(6[4-9]|[7-9][0-9]|1[0-1][0-9]|12[0-7])'
    map[2][128] = '(12[8-9]|1[3-8][0-9]|19[0-1])'
    map[2][192] = '(19[2-9]|2([0-4][0-9]|5[0-5]))'
    # 32
    map[3] = {}
    map[3][0] = '([0-9]|[1-2][0-9]|3[0-1])'
    map[3][32] = '(3[2-9]|[4-5][0-9]|6[0-3])'
    map[3][64] = '(6[4-9]|[7-8][0-9]|9[0-5])'
    map[3][96] = '(9[6-9]|1[0-1][0-9]|12[0-7])'
    map[3][128] = '(12[8-9]|1[3-5][0-9])'
    map[3][160] = '(1[6-8][0-9]|19[0-1])'
    map[3][192] = '(19[2-9]|2[0-1][0-9]|22[0-3])'
    map[3][224] = '(22[4-9]|2[3-4][0-9]|25[0-5])'
    # 16
    map[4] = {}
    map[4][0] = '([0-9]|1[0-5])'
    map[4][16] = '(1[6-9]|2[0-9]|3[0-1])'
    map[4][32] = '(3[2-9]|4[0-7])'
    map[4][48] = '(4[8-9]|5[0-9]|6[0-3])'
    map[4][64] = '(6[4-9]|7[0-9])'
    map[4][80] = '(8[0-9]|9[0-5])'
    map[4][96] = '(9[6-9]|10[0-9]|11[0-1])'
    map[4][112] = '(11[2-9]|12[0-7])'
    map[4][128] = '(12[8-9]|13[0-9]|14[0-3])'
    map[4][144] = '(14[4-9]|15[0-9])'
    map[4][160] = '(16[0-9]|17[0-5])'
    map[4][176] = '(17[6-9]|18[0-9]|19[0-1])'
    map[4][192] = '(19[2-9]|20[0-7])'
    map[4][208] = '(20[8-9]|21[0-9]|22[0-3])'
    map[4][224] = '(22[4-9]|23[0-9])'
    map[4][240] = '(24[0-9]|25[0-5])'
    # 8
    map[5] = {}
    map[5][0] = '[0-7]'
    map[5][8] = '([8-9]|1[0-5])'
    map[5][16] = '(1[6-9]|2[0-3])'
    map[5][24] = '(2[4-9]|3[0-1])'
    map[5][32] = '3[2-9]'
    map[5][40] = '4[0-7]'
    map[5][48] = '(4[8-9]|5[0-5])'
    map[5][56] = '(5[6-9]|6[0-3])'
    map[5][64] = '(6[4-9]|7[0-1])'
    map[5][72] = '7[2-9]'
    map[5][80] = '8[0-7]'
    map[5][88] = '(8[8-9]|9[0-5])'
    map[5][96] = '(9[6-9]|10[0-3])'
    map[5][104] = '(10[4-9]|11[0-1])'
    map[5][112] = '11[2-9]'
    map[5][120] = '12[0-7]'
    map[5][128] = '(12[8-9]|13[0-5])'
    map[5][136] = '(13[6-9]|14[0-3])'
    map[5][144] = '(14[4-9]|15[0-1])'
    map[5][152] = '15[2-9]'
    map[5][160] = '16[0-7]'
    map[5][168] = '(16[8-9]|17[0-5])'
    map[5][176] = '(17[6-9]|18[0-3])'
    map[5][184] = '(18[4-9]|19[0-1])'
    map[5][192] = '19[2-9]'
    map[5][200] = '20[0-7]'
    map[5][208] = '(20[8-9]|21[0-5])'
    map[5][216] = '(21[6-9]|22[0-3])'
    map[5][224] = '(22[4-9]|23[0-1])'
    map[5][232] = '23[2-9]'
    map[5][240] = '24[0-7]'
    map[5][248] = '(24[8-9]|25[0-5])'
    # 4
    map[6] = {}
    map[6][0] = '[0-3]'
    map[6][4] = '[4-7]'
    map[6][8] = '([8-9]|1[0-1])'
    map[6][12] = '1[2-5]'
    map[6][16] = '1[6-9]'
    map[6][20] = '2[0-3]'
    map[6][24] = '2[4-7]'
    map[6][28] = '(2[8-9]|3[0-1])'
    map[6][32] = '3[2-5]'
    map[6][36] = '3[6-9]'
    map[6][40] = '4[0-3]'
    map[6][44] = '4[4-7]'
    map[6][48] = '(4[8-9]|5[0-1])'
    map[6][52] = '5[2-5]'
    map[6][56] = '5[6-9]'
    map[6][60] = '6[0-3]'
    map[6][64] = '6[4-7]'
    map[6][68] = '(6[8-9]|7[0-1])'
    map[6][72] = '7[2-5]'
    map[6][76] = '7[6-9]'
    map[6][80] = '8[0-3]'
    map[6][84] = '8[4-7]'
    map[6][88] = '(8[8-9]|9[0-1])'
    map[6][92] = '9[2-5]'
    map[6][96] = '9[6-9]'
    map[6][100] = '10[0-3]'
    map[6][104] = '10[4-7]'
    map[6][108] = '(10[8-9]|11[0-1])'
    map[6][112] = '11[2-5]'
    map[6][116] = '11[6-9]'
    map[6][120] = '12[0-3]'
    map[6][124] = '12[4-7]'
    map[6][128] = '(12[8-9]|13[0-1])'
    map[6][132] = '13[2-5]'
    map[6][136] = '13[6-9]'
    map[6][140] = '14[0-3]'
    map[6][144] = '14[4-7]'
    map[6][148] = '(14[8-9]|15[0-1])'
    map[6][152] = '15[2-5]'
    map[6][156] = '15[6-9]'
    map[6][160] = '16[0-3]'
    map[6][164] = '16[4-7]'
    map[6][168] = '(16[8-9]|17[0-1])'
    map[6][172] = '17[2-5]'
    map[6][176] = '17[6-9]'
    map[6][180] = '18[0-3]'
    map[6][184] = '18[4-7]'
    map[6][188] = '(18[8-9]|19[0-1])'
    map[6][192] = '19[2-5]'
    map[6][196] = '19[6-9]'
    map[6][200] = '20[0-3]'
    map[6][204] = '20[4-7]'
    map[6][208] = '(20[8-9]|21[0-1])'
    map[6][212] = '21[2-5]'
    map[6][216] = '21[6-9]'
    map[6][220] = '22[0-3]'
    map[6][224] = '22[4-7]'
    map[6][228] = '(22[8-9]|23[0-1])'
    map[6][232] = '23[2-5]'
    map[6][236] = '23[6-9]'
    map[6][240] = '24[0-3]'
    map[6][244] = '24[4-7]'
    map[6][248] = '(24[8-9]|25[0-1])'
    map[6][252] = '25[2-5]'
    
    # Setup some vars
    network = cidr.split('/')[0]
    n1 = network.split('.')[0]
    n2 = network.split('.')[1]
    n3 = network.split('.')[2]
    n4 = network.split('.')[3]
    mask = cidr.split('/')[1]
    i_mask = int(mask)
    bip = 0
    #Seperator: in impala: '\\.' for single dot. in oother regex: '\.'
    sep = r'\\.'
    # Logic based on subnet mask
    if mask == '0':
        return map[0][0] + sep + map[0][0] + sep + map[0][0] + sep + map[0][0]
    elif 1 <= i_mask <= 6:
        pow = 2**(8-i_mask)
        bip = int(int(n1)/pow)*pow
        return map[i_mask][bip] + sep + map[0][0] + sep + map[0][0] + sep + map[0][0]
    elif mask == '7':
        bip = int(int(n1)/2)*2
        if len(str(bip)) == 2:
            return str(bip)[0] + '[' + str(bip)[1] + '-' + str(bip+1)[1] + ']' + sep + map[0][0] + sep + map[0][0] + sep + map[0][0]
        elif len(str(bip)) == 3:
            return str(bip)[:2] + '[' + str(bip)[2] + '-' + str(bip+1)[2] + ']' + sep + map[0][0] + sep + map[0][0] + sep + map[0][0]
        else:
            return '[' + str(bip) + '-' + str(bip+1) + ']' + sep + map[0][0] + sep + map[0][0] + sep + map[0][0]
        
    elif mask == '8':
        return n1 + sep + map[0][0] + sep + map[0][0] + sep + map[0][0]
    elif 9 <= i_mask <= 14:
        pow = 2**(16-i_mask)
        bip = int(int(n2)/pow)*pow
        return n1 + sep + map[i_mask-8][bip] + sep + map[0][0] + sep + map[0][0]
    elif mask == '15':
        bip = int(int(n2)/2)*2
        if len(str(bip)) == 2:
            return n1 + sep + str(bip)[0] + '[' + str(bip)[1] + '-' + str(bip+1)[1] + ']' + sep + map[0][0] + sep + map[0][0]
        elif len(str(bip)) == 3:
            return n1 + sep + str(bip)[:2] + '[' + str(bip)[2] + '-' + str(bip+1)[2] + ']' + sep + map[0][0] + sep + map[0][0]
        else:
            return n1 + sep + '[' + str(bip) + '-' + str(bip+1) + ']' + sep + map[0][0] + sep + map[0][0]
    
    elif mask == '16':
        return n1 + sep + n2 + sep + map[0][0] + sep + map[0][0]
    elif 17 <= i_mask <= 22:
        pow = 2**(24-i_mask)
        bip = int(int(n3)/pow)*pow
        return n1 + sep + n2 + sep + map[i_mask-16][bip] + sep + map[0][0]
    elif mask == '23':
        bip = int(int(n3)/2)*2
        if len(str(bip)) == 2:
            return n1 + sep + n2 + sep + str(bip)[0] + '[' + str(bip)[1] + '-' + str(bip+1)[1] + ']' + sep + map[0][0]
        elif len(str(bip)) == 3:
            return n1 + sep + n2 + sep + str(bip)[:2] + '[' + str(bip)[2] + '-' + str(bip+1)[2] + ']' + sep + map[0][0]
        else:
            return n1 + sep + n2 + sep + '[' + str(bip) + '-' + str(bip+1) + ']' + sep + map[0][0]
    
    elif mask == '24':
        return n1 + sep + n2 + sep + n3 + sep + map[0][0]
    elif 25 <= i_mask <= 30:
        pow = 2**(32-i_mask)
        bip = int(int(n4)/pow)*pow
        return n1 + sep + n2 + sep + n3 + sep + map[i_mask-24][bip]
    elif mask == '31':
        bip = int(int(n4)/2)*2
        if len(str(bip)) == 2:
            return n1 + sep + n2 + sep + n3 + sep + str(bip)[0] + '[' + str(bip)[1] + '-' + str(bip+1)[1] + ']'
        elif len(str(bip)) == 3:
            return n1 + sep + n2 + sep + n3 + sep + str(bip)[:2] + '[' + str(bip)[2] + '-' + str(bip+1)[2] + ']'
        else:
            return n1 + sep + n2 + sep + n3 + sep + '[' + str(bip) + '-' + str(bip+1) + ']'
    elif mask == '32':
        return n1 + sep + n2 + sep + n3 + sep + n4
    else:
        return 'Invalid Subnet Mask.'
        
 
 ######## section 2: #######  
 #Summarize a list of single and subnet IPs
 #Export unique signle IP list 
 #Export Summarized subnet list (CIDR) + REGEX
 ############################
def returnIPList(IPwithMask):
    ip = IPNetwork(IPwithMask)
    return list(ip)
def SummarizeIP(in_fn,out1_fn,out2_fn):
    # load pool IP addresses with netmask : sample : 10.0.96.0/19
    s = load_workbook(in_fn + '.xlsx')
    a = s.active
    
    ip_list = []
    # send all masked ip to returnIPList() method and get the list of its hosts
    for row in a.iter_rows():
        for cell in row:
            try:
                ip_list.append(returnIPList(cell.value))
            except Exception as ee:
                print(cell.value, ee)
            break
    #iplen = len(ip_list)
    #print(iplen)
    #write list of all uinique pool IP addresses to output file
    f = open(out1_fn + '.csv', 'w', newline='')
    writer = csv.writer(f, delimiter=',')
    all_ip_list = []
    for ip in ip_list:
        for i in ip:
            all_ip_list.append(i)
    unique_list = list(set(all_ip_list))
    unique_list.sort()
    print("Total Unique IPs:",len(unique_list))    
    for ip in unique_list:
         writer.writerow([ip])        
    f.close()
    # summarize and write summarized subnet IP (CIDR) and its REGEX to file
    summarized_list = IPSet(unique_list)
    f = open(out2_fn +'.csv', 'w', newline='')
    writer = csv.writer(f, delimiter=',')
    writer.writerow(['CIDR','REGEX'])
    subnet_list=[]
    for cidr in summarized_list.iter_cidrs():
        writer.writerow([cidr,"'^"+ cidr_to_regex(str(cidr))+"$'"])   
        subnet_list.append(cidr)
    f.close()
    return



######## section 3: #######
#Insert unique IP list to SQL file
###########################    
def replace_IP(ip_fn,sql_temp):
    sql_out =  sql_temp +"_"
    
    df0 = pd.read_csv(ip_fn + ".csv",header=None)
    df0.rename(columns={0: "IP"}, inplace=True)
    
    IP_list = list(df0['IP'].unique())
    IP_string = ""
    
    for ip in IP_list:
        IP_string += f"'{ip}', "
    
    query_template = open(query_path +'\\'+ f"{sql_temp}.sql", "r").read()
    
    f = open(query_path +'\\'+ f"{sql_out}.sql", 'w')
    
    f.write(query_template.replace('[?IP?]', IP_string[:-2]))
    f.close()  
    return



folder_IP = r"D:\...\Traffic_daily_IP"
query_path = folder_IP + r"\SQL_Templates"
sql_file = "ServerIP_Spark_Traffic"
IP_NET_file = 'IP_Net_pool'

chdir(folder_IP)    
file_IP_list = "UnqSingle_IP"
SummarizeIP (IP_NET_file,file_IP_list,'SubnetRegex')
replace_IP(file_IP_list, sql_file)

